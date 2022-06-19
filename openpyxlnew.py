import openpyxl as op

c= op.load_workbook("C:\\Users\\SUDHANSU\\Desktop\\a.xlsx")
print(c.sheetnames)
s = c["my"]
x=s.cell(row=1,column=2).value
print(x)
print(s['A2'].value)
print("multiple valuse using row")
print(s.max_row)
print(s.max_column)

d = s.max_column
for i in range(1,d+1):
    print(s.cell(row=2,column=i).value)

d1 = s.max_row
for i in range(1,d1+1):
    print(s.cell(row=i,column=2).value)
